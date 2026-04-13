"""
Automated Excel Report Email Dispatcher

This script simulates a freelance-grade reporting workflow where:
- recipient instructions are loaded from a CSV or Excel file in /input
- report files are collected from /input/reports
- rows are cleaned and validated
- matching report files are attached to recipient emails
- email delivery can run in DRY RUN mode or real SMTP mode
- a formatted Excel dispatch log is written to /output
- detailed logs are written to /log

Expected input files
--------------------
1) A recipient instruction file in /input
   Supported names:
   - recipient_distribution.xlsx
   - recipient_distribution.xls
   - recipient_distribution.csv
   - recipients.xlsx
   - recipients.csv

2) Report files inside:
   /input/reports/

Expected recipient columns
--------------------------
Required:
- recipient_name
- email_to
- report_file

Optional:
- company
- email_cc
- email_bcc
- subject
- message
- is_active

Examples:
recipient_name,email_to,report_file,subject,message,is_active
Alice Smith,alice@example.com,north_region_sales.xlsx,Monthly Sales Report,"Hi Alice, please find your report attached.",yes
Bob Jones,bob@example.com,finance_summary.xlsx,,,
"""

from __future__ import annotations

import logging
import mimetypes
import os
import re
import smtplib
import ssl
import sys
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# PATHS
# =============================================================================

BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "input"
REPORTS_DIR = INPUT_DIR / "reports"
OUTPUT_DIR = BASE_DIR / "output"
LOG_DIR = BASE_DIR / "log"

TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
LOG_FILE = LOG_DIR / f"report_dispatch_{TIMESTAMP}.log"
DISPATCH_LOG_FILE = OUTPUT_DIR / f"dispatch_register_{TIMESTAMP}.xlsx"
SUMMARY_FILE = OUTPUT_DIR / f"dispatch_summary_{TIMESTAMP}.txt"

# Safety default: do not send real emails unless explicitly enabled.
DRY_RUN = True

# Optional SMTP environment variable names
SMTP_HOST_ENV = "REPORT_SMTP_HOST"
SMTP_PORT_ENV = "REPORT_SMTP_PORT"
SMTP_USER_ENV = "REPORT_SMTP_USER"
SMTP_PASSWORD_ENV = "REPORT_SMTP_PASSWORD"
SMTP_SENDER_ENV = "REPORT_SMTP_SENDER"
ENABLE_SEND_ENV = "REPORT_ENABLE_SEND"


# =============================================================================
# LOGGING
# =============================================================================

def setup_directories() -> None:
    """Create required directories if they do not already exist."""
    for folder in (INPUT_DIR, REPORTS_DIR, OUTPUT_DIR, LOG_DIR):
        folder.mkdir(parents=True, exist_ok=True)


def setup_logging() -> None:
    """Configure file and console logging."""
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # Clear existing handlers in case script is re-run in the same interpreter.
    if logger.handlers:
        logger.handlers.clear()

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(funcName)s | %(message)s"
    )

    file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setLevel(logging.INFO)
    stream_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)


# =============================================================================
# FILE DISCOVERY
# =============================================================================

def find_recipient_file() -> Path:
    """
    Locate the main recipient distribution file inside /input.

    Returns:
        Path: The discovered recipient file.

    Raises:
        FileNotFoundError: If no supported file is found.
    """
    candidate_names = [
        "recipient_distribution.xlsx",
        "recipient_distribution.xls",
        "recipient_distribution.csv",
        "recipients.xlsx",
        "recipients.xls",
        "recipients.csv",
    ]

    for name in candidate_names:
        candidate = INPUT_DIR / name
        if candidate.exists():
            logging.info("Recipient file found: %s", candidate.name)
            return candidate

    supported = ", ".join(candidate_names)
    raise FileNotFoundError(
        f"No recipient input file found in '{INPUT_DIR}'. "
        f"Expected one of: {supported}"
    )


def load_table(file_path: Path) -> pd.DataFrame:
    """
    Load CSV or Excel data into a DataFrame.

    Args:
        file_path: Path to the source file.

    Returns:
        pd.DataFrame: Loaded data.
    """
    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        logging.info("Loading CSV input: %s", file_path.name)
        return pd.read_csv(file_path)

    if suffix in {".xlsx", ".xls"}:
        logging.info("Loading Excel input: %s", file_path.name)
        return pd.read_excel(file_path)

    raise ValueError(f"Unsupported file type: {file_path.suffix}")


def list_report_files() -> list[Path]:
    """
    Collect all report files from /input/reports.

    Returns:
        list[Path]: List of files.
    """
    if not REPORTS_DIR.exists():
        logging.warning("Report directory does not exist: %s", REPORTS_DIR)
        return []

    report_files = [path for path in REPORTS_DIR.iterdir() if path.is_file()]
    logging.info("Discovered %s report file(s) in %s", len(report_files), REPORTS_DIR)
    return report_files


# =============================================================================
# DATA CLEANING AND VALIDATION
# =============================================================================

def standardize_column_name(column_name: Any) -> str:
    """
    Convert raw column names to consistent snake_case style.
    """
    text = str(column_name).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Standardize DataFrame column names.
    """
    cleaned = df.copy()
    cleaned.columns = [standardize_column_name(col) for col in cleaned.columns]
    return cleaned


def normalize_text(value: Any) -> str:
    """
    Clean text values safely.
    """
    if pd.isna(value):
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_email_list(value: Any) -> str:
    """
    Normalize comma- or semicolon-separated email lists.
    """
    raw = normalize_text(value)
    if not raw:
        return ""

    separators_normalized = raw.replace(";", ",")
    emails = [email.strip() for email in separators_normalized.split(",") if email.strip()]
    return ",".join(emails)


def parse_yes_no(value: Any, default: bool = True) -> bool:
    """
    Convert a flexible truthy/falsy value into a boolean.
    """
    if pd.isna(value) or str(value).strip() == "":
        return default

    text = str(value).strip().lower()
    return text in {"yes", "y", "true", "1", "active", "on"}


def is_valid_email(address: str) -> bool:
    """
    Basic email format validation.
    """
    if not address:
        return False

    pattern = r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$"
    return bool(re.match(pattern, address))


def split_email_addresses(email_string: str) -> list[str]:
    """
    Split a normalized comma-separated email string into a list.
    """
    if not email_string:
        return []
    return [part.strip() for part in email_string.split(",") if part.strip()]


def clean_recipient_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean and normalize the recipient instruction DataFrame.

    Required columns:
    - recipient_name
    - email_to
    - report_file

    Optional columns:
    - company
    - email_cc
    - email_bcc
    - subject
    - message
    - is_active
    """
    cleaned = standardize_columns(df).copy()

    # Add missing optional columns for consistent downstream logic.
    optional_columns = [
        "company",
        "email_cc",
        "email_bcc",
        "subject",
        "message",
        "is_active",
    ]
    for column in optional_columns:
        if column not in cleaned.columns:
            cleaned[column] = ""

    required_columns = ["recipient_name", "email_to", "report_file"]
    missing = [col for col in required_columns if col not in cleaned.columns]
    if missing:
        raise ValueError(f"Missing required column(s): {missing}")

    # Drop fully empty rows first.
    original_rows = len(cleaned)
    cleaned = cleaned.dropna(how="all").copy()
    removed_empty = original_rows - len(cleaned)
    if removed_empty:
        logging.info("Removed %s fully empty row(s)", removed_empty)

    # Normalize text fields.
    text_columns = [
        "recipient_name",
        "email_to",
        "report_file",
        "company",
        "email_cc",
        "email_bcc",
        "subject",
        "message",
        "is_active",
    ]

    for col in text_columns:
        cleaned[col] = cleaned[col].apply(normalize_text)

    cleaned["email_to"] = cleaned["email_to"].apply(normalize_email_list)
    cleaned["email_cc"] = cleaned["email_cc"].apply(normalize_email_list)
    cleaned["email_bcc"] = cleaned["email_bcc"].apply(normalize_email_list)
    cleaned["is_active"] = cleaned["is_active"].apply(parse_yes_no)

    # Remove rows with blank critical fields.
    before_filter = len(cleaned)
    cleaned = cleaned[
        (cleaned["recipient_name"] != "")
        & (cleaned["email_to"] != "")
        & (cleaned["report_file"] != "")
    ].copy()
    removed_blank_critical = before_filter - len(cleaned)
    if removed_blank_critical:
        logging.info("Removed %s row(s) with blank critical fields", removed_blank_critical)

    # Remove duplicates based on delivery identity.
    before_dedup = len(cleaned)
    cleaned = cleaned.drop_duplicates(
        subset=["recipient_name", "email_to", "report_file"],
        keep="first",
    ).copy()
    removed_duplicates = before_dedup - len(cleaned)
    if removed_duplicates:
        logging.info("Removed %s duplicate row(s)", removed_duplicates)

    cleaned.reset_index(drop=True, inplace=True)
    return cleaned


def validate_recipient_data(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Split recipient data into valid and invalid records.

    Returns:
        tuple[pd.DataFrame, pd.DataFrame]: valid_df, invalid_df
    """
    records = df.copy()
    validation_issues: list[str] = []

    for _, row in records.iterrows():
        row_issues: list[str] = []

        to_addresses = split_email_addresses(row["email_to"])
        cc_addresses = split_email_addresses(row["email_cc"])
        bcc_addresses = split_email_addresses(row["email_bcc"])

        if not row["is_active"]:
            row_issues.append("inactive_recipient")

        if not to_addresses:
            row_issues.append("missing_to_email")

        invalid_to = [email for email in to_addresses if not is_valid_email(email)]
        invalid_cc = [email for email in cc_addresses if not is_valid_email(email)]
        invalid_bcc = [email for email in bcc_addresses if not is_valid_email(email)]

        if invalid_to:
            row_issues.append(f"invalid_to_email: {', '.join(invalid_to)}")
        if invalid_cc:
            row_issues.append(f"invalid_cc_email: {', '.join(invalid_cc)}")
        if invalid_bcc:
            row_issues.append(f"invalid_bcc_email: {', '.join(invalid_bcc)}")

        validation_issues.append(" | ".join(row_issues))

    records["validation_issues"] = validation_issues

    valid_df = records[records["validation_issues"] == ""].copy()
    invalid_df = records[records["validation_issues"] != ""].copy()

    logging.info("Valid rows: %s", len(valid_df))
    logging.info("Invalid rows: %s", len(invalid_df))
    return valid_df, invalid_df


# =============================================================================
# REPORT MATCHING
# =============================================================================

def build_report_lookup(report_files: list[Path]) -> dict[str, Path]:
    """
    Build a case-insensitive lookup of report filename to path.
    """
    lookup: dict[str, Path] = {}
    for file_path in report_files:
        lookup[file_path.name.lower()] = file_path
    return lookup


def resolve_report_path(report_reference: str, report_lookup: dict[str, Path]) -> Path | None:
    """
    Match a recipient's report_file value to an actual report file.

    Strategy:
    1) Exact filename match
    2) Case-insensitive filename match
    3) Stem match if extension omitted
    """
    ref = normalize_text(report_reference)
    if not ref:
        return None

    # Exact path by filename key
    exact = report_lookup.get(ref.lower())
    if exact:
        return exact

    # Try matching by stem if extension omitted
    ref_stem = Path(ref).stem.lower()
    for filename, path in report_lookup.items():
        if Path(filename).stem.lower() == ref_stem:
            return path

    return None


def attach_reports_to_recipients(
    recipients_df: pd.DataFrame,
    report_files: list[Path],
) -> pd.DataFrame:
    """
    Add matched report metadata to recipient rows.
    """
    result = recipients_df.copy()
    report_lookup = build_report_lookup(report_files)

    matched_paths: list[str] = []
    match_statuses: list[str] = []

    for _, row in result.iterrows():
        match = resolve_report_path(row["report_file"], report_lookup)

        if match is None:
            matched_paths.append("")
            match_statuses.append("report_not_found")
        else:
            matched_paths.append(str(match))
            match_statuses.append("matched")

    result["matched_report_path"] = matched_paths
    result["match_status"] = match_statuses
    return result


# =============================================================================
# EMAIL CONSTRUCTION
# =============================================================================

def get_smtp_settings() -> dict[str, Any]:
    """
    Read SMTP settings from environment variables.

    Required for real email sending:
    - REPORT_SMTP_HOST
    - REPORT_SMTP_PORT
    - REPORT_SMTP_USER
    - REPORT_SMTP_PASSWORD
    - REPORT_SMTP_SENDER

    Optional:
    - REPORT_ENABLE_SEND=1 to disable dry run
    """
    enable_send = os.getenv(ENABLE_SEND_ENV, "").strip() == "1"

    settings = {
        "host": os.getenv(SMTP_HOST_ENV, "").strip(),
        "port": os.getenv(SMTP_PORT_ENV, "").strip(),
        "user": os.getenv(SMTP_USER_ENV, "").strip(),
        "password": os.getenv(SMTP_PASSWORD_ENV, "").strip(),
        "sender": os.getenv(SMTP_SENDER_ENV, "").strip(),
        "enable_send": enable_send,
    }
    return settings


def derive_subject(row: pd.Series) -> str:
    """
    Build email subject, using the input value if present.
    """
    custom_subject = normalize_text(row.get("subject", ""))
    if custom_subject:
        return custom_subject

    report_name = Path(row["report_file"]).stem.replace("_", " ").replace("-", " ").title()
    company = normalize_text(row.get("company", ""))
    if company:
        return f"{company} | {report_name}"
    return f"Automated Report Delivery | {report_name}"


def derive_message_body(row: pd.Series) -> str:
    """
    Build email body, using the input value if present.
    """
    custom_message = normalize_text(row.get("message", ""))
    if custom_message:
        return custom_message

    recipient_name = normalize_text(row["recipient_name"])
    report_name = Path(row["report_file"]).name

    return (
        f"Hello {recipient_name},\n\n"
        f"Please find attached your requested report: {report_name}.\n\n"
        f"This message was generated automatically by the reporting workflow.\n\n"
        f"Best regards,\n"
        f"Reporting Automation System"
    )


def add_attachment(message: EmailMessage, file_path: Path) -> None:
    """
    Attach a file to an email message.
    """
    mime_type, _ = mimetypes.guess_type(file_path.name)
    if mime_type:
        maintype, subtype = mime_type.split("/", 1)
    else:
        maintype, subtype = "application", "octet-stream"

    with open(file_path, "rb") as file:
        message.add_attachment(
            file.read(),
            maintype=maintype,
            subtype=subtype,
            filename=file_path.name,
        )


def build_email_message(row: pd.Series, sender_email: str) -> EmailMessage:
    """
    Create a fully formed email message for a recipient row.
    """
    message = EmailMessage()
    message["From"] = sender_email
    message["To"] = row["email_to"]

    if normalize_text(row.get("email_cc", "")):
        message["Cc"] = row["email_cc"]

    if normalize_text(row.get("email_bcc", "")):
        message["Bcc"] = row["email_bcc"]

    message["Subject"] = derive_subject(row)
    message.set_content(derive_message_body(row))

    report_path = Path(row["matched_report_path"])
    add_attachment(message, report_path)

    return message


# =============================================================================
# EMAIL DISPATCH
# =============================================================================

def send_email_message(message: EmailMessage, smtp_settings: dict[str, Any]) -> None:
    """
    Send a single email via SMTP with TLS.
    """
    host = smtp_settings["host"]
    port = int(smtp_settings["port"])
    user = smtp_settings["user"]
    password = smtp_settings["password"]

    context = ssl.create_default_context()

    with smtplib.SMTP(host, port) as server:
        server.starttls(context=context)
        server.login(user, password)
        server.send_message(message)


def dispatch_emails(dispatch_df: pd.DataFrame) -> pd.DataFrame:
    """
    Process all matched recipient rows for delivery.

    Real sending only occurs when REPORT_ENABLE_SEND=1 and
    all required SMTP environment variables exist.
    """
    smtp_settings = get_smtp_settings()

    global DRY_RUN
    DRY_RUN = not smtp_settings["enable_send"]

    required_smtp_values = [
        smtp_settings["host"],
        smtp_settings["port"],
        smtp_settings["user"],
        smtp_settings["password"],
        smtp_settings["sender"],
    ]
    smtp_ready = all(required_smtp_values)

    if not DRY_RUN and not smtp_ready:
        logging.warning(
            "Real sending was enabled, but SMTP settings are incomplete. "
            "Falling back to DRY RUN mode."
        )
        DRY_RUN = True

    logging.info("Dispatch mode: %s", "DRY RUN" if DRY_RUN else "LIVE SEND")

    results: list[dict[str, Any]] = []

    for _, row in dispatch_df.iterrows():
        record = row.to_dict()

        # Default outcome values
        delivery_status = "not_attempted"
        delivery_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        error_message = ""

        try:
            if row["match_status"] != "matched":
                delivery_status = "failed"
                error_message = "Report file not matched"
                logging.warning(
                    "Skipping %s because the report could not be matched",
                    row["recipient_name"],
                )
            else:
                sender_email = smtp_settings["sender"] or "no-reply@example.com"
                email_message = build_email_message(row, sender_email)

                if DRY_RUN:
                    delivery_status = "dry_run_success"
                    logging.info(
                        "DRY RUN email prepared for %s -> %s | attachment=%s",
                        row["recipient_name"],
                        row["email_to"],
                        Path(row["matched_report_path"]).name,
                    )
                else:
                    send_email_message(email_message, smtp_settings)
                    delivery_status = "sent"
                    logging.info(
                        "Email sent successfully to %s -> %s",
                        row["recipient_name"],
                        row["email_to"],
                    )

        except Exception as exc:  # noqa: BLE001
            delivery_status = "failed"
            error_message = str(exc)
            logging.exception(
                "Failed to process email for %s (%s)",
                row.get("recipient_name", ""),
                row.get("email_to", ""),
            )

        record["delivery_status"] = delivery_status
        record["delivery_timestamp"] = delivery_timestamp
        record["error_message"] = error_message
        results.append(record)

    return pd.DataFrame(results)


# =============================================================================
# OUTPUT REPORTING
# =============================================================================

def autosize_worksheet_columns(worksheet) -> None:
    """
    Auto-fit worksheet columns based on content length.
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column_index = column_cells[0].column
        column_letter = get_column_letter(column_index)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        worksheet.column_dimensions[column_letter].width = min(max_length + 3, 60)


def apply_excel_styling(file_path: Path) -> None:
    """
    Apply professional formatting to the dispatch Excel register.
    """
    workbook = load_workbook(file_path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    status_fills = {
        "sent": PatternFill("solid", fgColor="C6EFCE"),
        "dry_run_success": PatternFill("solid", fgColor="DDEBF7"),
        "failed": PatternFill("solid", fgColor="F4CCCC"),
        "not_attempted": PatternFill("solid", fgColor="FFF2CC"),
    }

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Highlight delivery status column when present
        headers = [cell.value for cell in worksheet[1]]
        if "delivery_status" in headers:
            status_col_index = headers.index("delivery_status") + 1
            for row_num in range(2, worksheet.max_row + 1):
                status_cell = worksheet.cell(row=row_num, column=status_col_index)
                status_value = str(status_cell.value).strip().lower()
                fill = status_fills.get(status_value)
                if fill:
                    status_cell.fill = fill

        autosize_worksheet_columns(worksheet)

    workbook.save(file_path)


def write_dispatch_outputs(
    valid_df: pd.DataFrame,
    invalid_df: pd.DataFrame,
    dispatch_results_df: pd.DataFrame,
) -> None:
    """
    Write Excel and text output files.
    """
    with pd.ExcelWriter(DISPATCH_LOG_FILE, engine="openpyxl") as writer:
        dispatch_results_df.to_excel(writer, sheet_name="dispatch_results", index=False)
        valid_df.to_excel(writer, sheet_name="valid_recipients", index=False)
        invalid_df.to_excel(writer, sheet_name="invalid_recipients", index=False)

    apply_excel_styling(DISPATCH_LOG_FILE)

    sent_count = int((dispatch_results_df["delivery_status"] == "sent").sum()) if not dispatch_results_df.empty else 0
    dry_run_count = int((dispatch_results_df["delivery_status"] == "dry_run_success").sum()) if not dispatch_results_df.empty else 0
    failed_count = int((dispatch_results_df["delivery_status"] == "failed").sum()) if not dispatch_results_df.empty else 0
    matched_count = int((dispatch_results_df["match_status"] == "matched").sum()) if not dispatch_results_df.empty else 0

    summary_lines = [
        "AUTOMATED EXCEL REPORT EMAIL DISPATCH SUMMARY",
        "=" * 55,
        f"Run timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Base directory: {BASE_DIR}",
        f"Input directory: {INPUT_DIR}",
        f"Reports directory: {REPORTS_DIR}",
        f"Output file: {DISPATCH_LOG_FILE.name}",
        f"Log file: {LOG_FILE.name}",
        f"Dispatch mode: {'DRY RUN' if DRY_RUN else 'LIVE SEND'}",
        "",
        "COUNTS",
        "-" * 55,
        f"Total valid recipients: {len(valid_df)}",
        f"Total invalid recipients: {len(invalid_df)}",
        f"Matched reports: {matched_count}",
        f"Emails sent: {sent_count}",
        f"Dry run successes: {dry_run_count}",
        f"Failures: {failed_count}",
        "",
        "NOTES",
        "-" * 55,
        "Real email sending requires these environment variables:",
        f"- {SMTP_HOST_ENV}",
        f"- {SMTP_PORT_ENV}",
        f"- {SMTP_USER_ENV}",
        f"- {SMTP_PASSWORD_ENV}",
        f"- {SMTP_SENDER_ENV}",
        f"- {ENABLE_SEND_ENV}=1",
    ]

    SUMMARY_FILE.write_text("\n".join(summary_lines), encoding="utf-8")

    logging.info("Dispatch register written to: %s", DISPATCH_LOG_FILE)
    logging.info("Summary report written to: %s", SUMMARY_FILE)


# =============================================================================
# MAIN WORKFLOW
# =============================================================================

def run_pipeline() -> None:
    """
    Execute the full report email distribution workflow.
    """
    logging.info("Starting Automated Excel Report Email Dispatcher")
    logging.info("Base directory: %s", BASE_DIR)

    recipient_file = find_recipient_file()
    recipient_df = load_table(recipient_file)

    cleaned_recipients = clean_recipient_data(recipient_df)
    valid_recipients, invalid_recipients = validate_recipient_data(cleaned_recipients)

    report_files = list_report_files()
    if not report_files:
        logging.warning("No report files found in %s", REPORTS_DIR)

    valid_recipients = attach_reports_to_recipients(valid_recipients, report_files)
    dispatch_results = dispatch_emails(valid_recipients)

    write_dispatch_outputs(
        valid_df=valid_recipients,
        invalid_df=invalid_recipients,
        dispatch_results_df=dispatch_results,
    )

    logging.info("Workflow completed successfully")


def main() -> None:
    """
    Script entry point with top-level error handling.
    """
    setup_directories()
    setup_logging()

    try:
        run_pipeline()
    except Exception as exc:  # noqa: BLE001
        logging.exception("Fatal workflow error: %s", exc)
        raise


if __name__ == "__main__":
    main()